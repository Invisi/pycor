import datetime
import logging
import os
import zipfile
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl.worksheet.worksheet
import openpyxl.reader.excel
import pywintypes
import typing
import win32com.client
from cryptography import fernet

import utils

config = utils.import_config()


class ExcelFileException(Exception):
    pass


def setup_excel():
    excel = win32com.client.Dispatch("Excel.Application")

    # "Do you want to save your work?"
    excel.DisplayAlerts = False
    # Links = Copied values from another sheet, we might want to detect those
    excel.AskToUpdateLinks = False
    # Speed up macro access
    excel.ScreenUpdating = False

    if config.SHOW_EXCEL:
        excel.Visible = True
        excel.ScreenUpdating = True

    return excel


class Commons:
    def __init__(self, excel_file: Path):
        self.excel_file = excel_file
        self.parent_path = Path(os.path.abspath(excel_file.parent))

        self.log = logging.getLogger("PyCor").getChild("Excel")
        self.log.info("Opening %s", self.get_relevant_path())

        self.exercise_ranges = None
        self.solutions = None

    def get_relevant_path(self):
        return os.sep.join(self.excel_file.parts[-3:])

    def set_exercise_rows(self, get_cell: typing.Callable, is_student: bool = False):
        """
        Extract amount of exercises, their ranges, and (if needed) entered solutions

        :param get_cell: Function to access the cells
        :param is_student: Whether to grab the solutions or not, only necessary in student files
        :return:
        """

        offset = 13
        index_num = offset
        previous_exercise = 0
        exercise_row_begin = []
        exercise_row_end = []
        self.exercise_ranges = []
        if is_student:
            self.solutions = []

        while True:
            current_exercise = get_cell(index_num, 1)  # A{index_num}

            # It's None once we are past all listed exercises or once we hit a hole
            if current_exercise is None:
                exercise_row_end.append(index_num - 1)
                break

            # Parse as int, convention
            try:
                current_exercise = int(current_exercise)
                if current_exercise <= 0:
                    raise ValueError
            except ValueError:
                raise ExcelFileException(
                    f"Failed to parse exercise number in A{index_num}"
                )

            # We hit a new exercise
            if current_exercise > previous_exercise:
                if previous_exercise > 0:
                    exercise_row_end.append(index_num - 1)
                exercise_row_begin.append(index_num)

            if is_student:
                # Collect submitted solutions
                if len(self.solutions) <= current_exercise - 1:
                    self.solutions.append([])

                self.solutions[current_exercise - 1].append(
                    get_cell(index_num, 3)
                )  # C{index_sum}
            else:
                # Verify tolerances are set correctly
                try:
                    tolerance_rel = get_cell(index_num, 4)  # D{index}
                    if tolerance_rel:
                        _ = float(tolerance_rel)
                except ValueError:
                    utils.write_error(
                        self.parent_path,
                        f"Ungültige relative Toleranz in D{index_num}.",
                    )
                    raise ExcelFileException("Invalid relative tolerance.")

                try:
                    tolerance_abs = get_cell(index_num, 5)  # E{index}
                    if tolerance_abs:
                        _ = float(tolerance_abs)
                except ValueError:
                    utils.write_error(
                        self.parent_path,
                        f"Ungültige absolute Toleranz in E{index_num}.",
                    )
                    raise ExcelFileException("Invalid absolute tolerance.")

            previous_exercise = current_exercise
            index_num += +1

        for i in range(len(exercise_row_begin)):
            self.exercise_ranges.append([exercise_row_begin[i], exercise_row_end[i]])

    def __str__(self):
        return self.excel_file


class Student(Commons):
    def __init__(self, excel_file: Path):
        super().__init__(excel_file)

        self.student_email = self.parent_path.name

        wb = None
        try:
            # Whether the file is considered valid
            self.valid = False

            # Open workbook (read-only) and grab first worksheet
            # Ignore formulas, ignore Excel's "smart" types
            wb = openpyxl.load_workbook(self.excel_file, read_only=True, data_only=True)
            ws = wb.worksheets[0]

            self.mat_num = int(ws.cell(10, 2).value or -1)
            self.dummies = [
                _.value
                for _ in [ws.cell(9, column) for column in range(2, 10)]  # B9 - I9
            ]

            if self.mat_num < 0:
                self.log.error(
                    "Invalid matriculation number specified in student file."
                )
                raise ExcelFileException("Invalid mat_num")

            self.set_exercise_rows(
                is_student=True, get_cell=lambda r, c: ws.cell(r, c).value
            )
            self.valid = True
        except (TypeError, ValueError):
            self.log.exception("Failed to read info from student file.")
            raise ExcelFileException("Failed to read information from student file.")
        except zipfile.BadZipFile:
            self.log.exception("Failed to open file, maybe it's password-protected?")
            raise ExcelFileException("Password-protected file?")
        finally:
            if wb:
                wb.close()

    def get_stats(self, exercise: int, max_attempts: int) -> tuple:
        """
        Returns the student's statistics

        :param exercise: Exercise number [beginning at 0]
        :param max_attempts: Maximum amount of tries before being blocked
        :return:
        """
        try:
            exercise_file = self.parent_path / "Exercise{}_block.txt".format(
                exercise + 1
            )

            if exercise_file.exists():
                # noinspection PyTypeChecker
                block_status = np.loadtxt(exercise_file)

                # Check if user's try list doesn't match the specified max_tries
                if len(block_status) > max_attempts:
                    # TODO: Shorten list and grab all attempts >0 in correct order
                    block_status = block_status[0:max_attempts]
                    np.savetxt(exercise_file, block_status, fmt="%3.2f")
                elif len(block_status) < max_attempts:
                    # Extend list
                    block_status = block_status + [0] * (
                        max_attempts - len(block_status)
                    )
                    np.savetxt(exercise_file, block_status, fmt="%3.2f")

                if 0 < block_status[-1] < 100:
                    # Last entry isn't passed
                    return True, False
                elif 100 in block_status:
                    # Any entry is marked as passed
                    return False, True
                else:
                    # Neither passed nor blocked
                    return False, False
            else:
                return False, False
        except IOError:
            self.log.exception("Failed to get student's stats.")
            raise

    def update_stats(
        self, exercise: int, correct_percentage: int, max_attempts: int
    ) -> tuple:
        """
        Updates and saves the student's statistics

        :param exercise: Exercise number [beginning at 0]
        :param correct_percentage: Percentage of correctly answered sub tasks
        :param max_attempts: Maximum amount of tries before being blocked
        :return:
        """
        try:
            blocked = False
            passed = correct_percentage == 100

            exercise_file = self.parent_path / "Exercise{}_block.txt".format(
                exercise + 1
            )
            student_data = self.parent_path / "data"

            # Does the file exist already?
            if exercise_file.exists():
                # noinspection PyTypeChecker
                block_status = np.loadtxt(exercise_file)
            else:
                block_status = np.zeros(max_attempts)

            # Check if user still has tries
            try_row = [r[0] for r in enumerate(block_status) if r[1] == 0]

            # There's at least one attempt left, let's log the results!
            if len(try_row) > 0:
                block_status[try_row[0]] = correct_percentage
                np.savetxt(exercise_file, block_status, fmt="%3.2f")
            else:
                blocked = True

            # Check if student failed his last try
            if 0 < block_status[-1] < 100:
                blocked = True
                self.log.info(
                    "Blocked %s for exercise %s", self.student_email, exercise + 1
                )

            # Create user-specific data folder
            if not student_data.exists():
                student_data.mkdir(exist_ok=True)

            current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Save percentage and date/time to file
            exercise_all = student_data / "Exercise{}.txt".format(exercise + 1)
            with exercise_all.open("a") as ea:
                ea.write("{} - {}\n".format(current_datetime, correct_percentage))

            # Save used mat_num to file
            mat_num_file = student_data / "mat_num.txt"
            with mat_num_file.open("a") as mnf:
                mnf.write("{} - {}\n".format(current_datetime, self.mat_num))

            return blocked, passed
        except IOError:
            self.log.exception("Failed to save student's stats.")
            raise


class Corrector(Commons):
    def __init__(self, excel_file: Path, simple_validation=True):
        super().__init__(excel_file)
        self.password = self.find_password()

        # Password could not be decrypted
        if self.password is None:
            self.valid = False
            return

        wb = None
        excel = None
        try:
            # Whether the file is considered valid
            self.valid = False

            if simple_validation and self.password == "":
                if self.excel_file.suffix == ".xls":  # openpyxl only handles xlsx/xlsm
                    self.log.warning("Found .xls file, trying to convert.")
                    self.convert_to_xlsx()
                    return
                else:
                    wb = openpyxl.load_workbook(
                        self.excel_file, read_only=True, data_only=True
                    )
                    ws = wb.worksheets[0]
            else:
                self.log.debug("File requires a password, can't open without Excel")
                """
                   Open workbook in Excel. It has to be Excel because workbook-wide encryption creates a weird FAT-like 
                   compound archive that can't be read with any (currently) existing library.
                """
                excel = setup_excel()
                wb = excel.Workbooks.Open(
                    self.excel_file, 0, False, None, self.password
                )
                ws = wb.Worksheets(1)

            def get_cell(row, column):
                if isinstance(ws, openpyxl.reader.excel.ReadOnlyWorksheet):
                    return ws.cell(row, column).value
                else:
                    return ws.Cells(row, column).Value

            # Extract subject
            self.corrector_title = get_cell(1, 2)  # B1
            if not self.corrector_title:
                utils.write_error(self.parent_path, "Ungültiger Name im Titel.")
                raise ExcelFileException(
                    "Empty title field. Please specify a valid name."
                )

            # Name that should be matched against submitted files
            self.codename = get_cell(2, 2)  # B2
            if not self.codename:
                utils.write_error(
                    self.parent_path, "Dateiname konnte nicht ausgelesen werden."
                )
                raise ExcelFileException(
                    "Empty file name field. Please specify a valid name."
                )
            elif ".xlsx" in self.codename:
                # Remove file ending, might get added accidentally
                self.codename = self.codename.replace(".xlsx", "")

            self.codename = str(self.codename).strip()

            # Check deadline, allow for same-day submissions
            self.deadline: datetime.datetime = get_cell(3, 2)  # B3
            if not isinstance(self.deadline, datetime.datetime):
                utils.write_error(self.parent_path, "Ungültige Frist.")
                raise ExcelFileException("Invalid deadline.")

            self.deadline = self.deadline.date()

            if (self.deadline - datetime.date.today()).days < 0:
                self.log.info(
                    "Ignoring %s due to deadline (%s)",
                    self.get_relevant_path(),
                    self.deadline,
                )
                return

            # Get max amount of attempts
            self.max_attempts = int(get_cell(4, 2) or 0)  # B4
            if self.max_attempts < 1:
                utils.write_error(self.parent_path, "Ungültige Anzahl an Versuchen.")
                raise ExcelFileException("Invalid number of max attempts.")

            # Grab exercise info
            self.set_exercise_rows(get_cell=get_cell)

            self.valid = True
        except (pywintypes.com_error, TypeError, ValueError):
            self.log.exception("Failed to read information from corrector.")
            utils.write_error(
                self.parent_path, "Fehler beim Einlesen der corrector-Datei."
            )
        except ExcelFileException:
            self.log.exception("Error in corrector file.")
        except AttributeError:
            self.log.exception("Looks like excel crashed. Quitting.")
            raise
        finally:
            # Close WorkBook and Excel
            if excel:
                if wb:
                    wb.Close(SaveChanges=False)
                excel.Application.Quit()
                del excel
            else:
                if wb and isinstance(wb, openpyxl.worksheet.worksheet.Worksheet):
                    wb.close()

    def generate_solutions(self, mat_num: int, dummies: []) -> Optional[list]:
        """
        Generates mat_num- and dummy-specific solutions and returns them in a two-dimensional list
        with the first dimension being the exercise and the second one containing
        the solution's name, value, and tolerance.

        :param mat_num: Student's matriculation number
        :param dummies: List of dummy values (a1-a7)
        :return:
        """
        wb = None
        excel = None

        try:
            excel = setup_excel()

            # Open workbook
            wb = excel.Workbooks.Open(self.excel_file, 0, False, None, self.password)
            ws = wb.Worksheets(1)

            # Copy values
            ws.Range("B10").Value = mat_num
            ws.Range("B9:I9").Value = dummies

            # Collect solutions
            solutions = []
            for idx, exercise in enumerate(self.exercise_ranges):
                if len(solutions) <= idx:
                    solutions.append([])

                for cell_number in range(exercise[0], exercise[1] + 1):
                    solutions[idx].append(
                        {
                            "name": ws.Cells(cell_number, 2).Value,  # B{index}
                            "value": ws.Cells(cell_number, 3).Value,  # C{index}
                            "tolerance_rel": ws.Cells(cell_number, 4).Value,  # D{index}
                            "tolerance_abs": ws.Cells(cell_number, 5).Value,  # E
                        }
                    )

            return solutions
        except (pywintypes.com_error, TypeError, ValueError):
            self.log.exception("Failed to generate solutions in corrector.")
            raise ExcelFileException("Failed to generate solutions.")
        except AttributeError:
            self.log.exception("Looks like excel crashed. Quitting.")
            raise
        finally:
            # Close WorkBook and Excel
            if wb:
                wb.Close(SaveChanges=False)
            if excel:
                excel.Application.Quit()
                del excel

    def convert_to_xlsx(self):
        """
        Convert current Excel file to .xlsx for openpyxl
        :return:
        """
        if self.excel_file.with_suffix(".xlsx").exists():
            self.log.warning(".xlsx already exists")
            self.excel_file.rename(
                self.excel_file.with_name(self.excel_file.name + "_konvertiert")
            )
            return

        excel = None
        wb = None
        try:
            excel = setup_excel()
            wb = excel.Workbooks.Open(self.excel_file, 0, False, None, self.password)
            wb.SaveAs(
                str(self.excel_file.with_suffix(".xlsx")), FileFormat=51
            )  # 51 = .xlsx
            wb.Close(SaveChanges=False)

            # Rename old file
            self.excel_file.rename(
                self.excel_file.with_name(self.excel_file.name + "_konvertiert")
            )
        except (pywintypes.com_error, TypeError, ValueError, AttributeError):
            self.log.exception("Failed to convert Excel file.")
            utils.write_error(
                self.parent_path,
                "Fehler beim Konvertieren der Corrector-Datei. Bitte speichern Sie sie als .xlsx/.xlsm.",
            )
        finally:
            if excel:
                excel.Application.Quit()
                del excel

    @staticmethod
    def from_path(subject_folder: Path):
        """
        Searches for `corrector.xls[mx]?` in given path and returns :class:`Corrector` if found

        :param subject_folder: Path in which should be searched
        """
        extensions = [".xlsx", ".xlsm", ".xls"]

        # Ignore folders containing the ignore file
        ignore_file = subject_folder / "PYCOR_IGNORE.txt"
        if ignore_file.exists():
            return
        for item in subject_folder.iterdir():
            if (
                item.is_file()
                and item.suffix in extensions
                and item.stem == "corrector"
            ):
                return Corrector(item)

    def find_password(self) -> Optional[str]:
        # Look for psw file
        try:
            psw_file = self.parent_path / "psw"
            if psw_file.exists():
                return (
                    fernet.Fernet(config.PSW_PASSPHRASE)
                    .decrypt(psw_file.read_bytes())
                    .decode("utf-8")
                )
            return ""
        except (
            fernet.InvalidToken,
            IOError,
        ):  # Decryption failed, ignore corrector file
            self.log.exception("Failed to decrypt psw for %s", self.get_relevant_path())
            utils.write_error(
                self.parent_path, "Fehler beim Entschlüsseln der Passwortdatei."
            )
            return None
